#!/usr/bin/env python3
# ================================================================
#  SIEM Africa — Module 6 : Generateur de rapports
#  Fichier  : reports/report_generator.py
#  Version  : 2.0
#
#  Types : incident, hebdomadaire, trimestriel, annuel, manuel
#  Formats : PDF (ReportLab) + Excel (OpenPyXL)
# ================================================================

import os
import sys
import sqlite3
import datetime
import logging

# ── Imports conditionnels ─────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
        Table, TableStyle, PageBreak, HRFlowable)
    from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER, TA_LEFT
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
        Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── Configuration ─────────────────────────────────────────────────
ENV_FILE    = '/opt/siem-africa/.env'
REPORTS_DIR = '/opt/siem-africa/rapports'

def load_env():
    cfg = {
        'DB_PATH':      '/opt/siem-africa/siem_africa.db',
        'REPORTS_DIR':  '/opt/siem-africa/rapports',
        'ORG_NOM':      'Mon Entreprise',
        'LANG':         'fr',
    }
    if os.path.exists(ENV_FILE):
        with open(ENV_FILE) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, _, v = line.partition('=')
                    cfg[k.strip()] = v.strip().strip('"').strip("'")
    return cfg

log = logging.getLogger('siem-reports')

# ── Couleurs PDF ──────────────────────────────────────────────────
NOIR    = HexColor('#111111') if PDF_OK else None
SUB     = HexColor('#333333') if PDF_OK else None
MUTED   = HexColor('#666666') if PDF_OK else None
BGT     = HexColor('#F0F0F0') if PDF_OK else None
BGL     = HexColor('#F8F8F8') if PDF_OK else None
BORD    = HexColor('#CCCCCC') if PDF_OK else None
TH_BG   = HexColor('#1A1A2E') if PDF_OK else None
CRIT    = HexColor('#C0392B') if PDF_OK else None
HIGH    = HexColor('#E67E22') if PDF_OK else None
MED     = HexColor('#F1C40F') if PDF_OK else None
LOW     = HexColor('#27AE60') if PDF_OK else None

# ── Base de donnees ───────────────────────────────────────────────
class DB:
    def __init__(self, db_path):
        self.path = db_path

    def conn(self):
        c = sqlite3.connect(self.path, timeout=10)
        c.row_factory = sqlite3.Row
        return c

    def get_param(self, cle, defaut=''):
        with self.conn() as c:
            r = c.execute('SELECT valeur FROM parametres WHERE cle=?', (cle,)).fetchone()
            return r['valeur'] if r else defaut

    # ── Donnees pour les rapports ──────────────────────────────────
    def alertes_periode(self, debut, fin):
        with self.conn() as c:
            return [dict(r) for r in c.execute('''
                SELECT a.*, at.nom as attaque_nom, at.mitre_id, at.mitre_tactique, at.categorie
                FROM alertes a
                LEFT JOIN attaques at ON a.attaque_id = at.id
                WHERE a.timestamp_alerte BETWEEN ? AND ?
                ORDER BY a.timestamp_alerte DESC
            ''', (debut, fin)).fetchall()]

    def stats_periode(self, debut, fin):
        with self.conn() as c:
            def q(sql, p=()): return c.execute(sql, p).fetchone()[0] or 0
            return {
                'total':      q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ?", (debut,fin)),
                'critiques':  q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ? AND gravite=4", (debut,fin)),
                'hautes':     q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ? AND gravite=3", (debut,fin)),
                'moyennes':   q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ? AND gravite=2", (debut,fin)),
                'faibles':    q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ? AND gravite=1", (debut,fin)),
                'resolues':   q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ? AND statut='Resolu'", (debut,fin)),
                'faux_pos':   q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ? AND statut='Faux positif'", (debut,fin)),
                'honeypot':   q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ? AND est_honeypot=1", (debut,fin)),
                'correlees':  q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ? AND est_correllee=1", (debut,fin)),
                'inconnues':  q("SELECT COUNT(*) FROM alertes WHERE timestamp_alerte BETWEEN ? AND ? AND est_inconnue=1", (debut,fin)),
                'ips_bloquees': q("SELECT COUNT(*) FROM ips_bloquees WHERE bloque_le BETWEEN ? AND ? AND type_blocage!='Whitelist'", (debut,fin)),
            }

    def top_attaques(self, debut, fin, limit=10):
        with self.conn() as c:
            return [dict(r) for r in c.execute('''
                SELECT nom_attaque, categorie, COUNT(*) as nb,
                       MAX(gravite) as gravite_max,
                       COUNT(DISTINCT ip_source) as nb_ips
                FROM alertes
                WHERE timestamp_alerte BETWEEN ? AND ?
                GROUP BY nom_attaque ORDER BY nb DESC LIMIT ?
            ''', (debut, fin, limit)).fetchall()]

    def top_ips(self, debut, fin, limit=10):
        with self.conn() as c:
            return [dict(r) for r in c.execute('''
                SELECT ip_source, pays_source, COUNT(*) as nb,
                       MAX(gravite) as gravite_max
                FROM alertes
                WHERE timestamp_alerte BETWEEN ? AND ?
                AND ip_source IS NOT NULL
                GROUP BY ip_source ORDER BY nb DESC LIMIT ?
            ''', (debut, fin, limit)).fetchall()]

    def alertes_par_jour(self, debut, fin):
        with self.conn() as c:
            return [dict(r) for r in c.execute('''
                SELECT date(timestamp_alerte) as jour, COUNT(*) as nb,
                       SUM(CASE WHEN gravite=4 THEN 1 ELSE 0 END) as critiques
                FROM alertes
                WHERE timestamp_alerte BETWEEN ? AND ?
                GROUP BY jour ORDER BY jour
            ''', (debut, fin)).fetchall()]

    def get_alerte(self, alerte_id):
        with self.conn() as c:
            return dict(c.execute('''
                SELECT a.*, at.nom as attaque_nom, at.mitre_id, at.mitre_tactique,
                       at.action_recommandee, at.contre_mesure, at.categorie,
                       u.email as resolu_par_email
                FROM alertes a
                LEFT JOIN attaques at ON a.attaque_id = at.id
                LEFT JOIN utilisateurs u ON a.resolu_par = u.id
                WHERE a.id = ?
            ''', (alerte_id,)).fetchone() or {})

    def sauver_rapport(self, type_rapport, fichier_pdf, fichier_excel,
                       debut, fin, nb_alertes, org):
        with self.conn() as c:
            c.execute('''
                INSERT INTO rapports
                (type_rapport, fichier_pdf, fichier_excel, periode_debut,
                 periode_fin, nb_alertes, organisation, cree_le)
                VALUES (?,?,?,?,?,?,?,?)
            ''', (type_rapport, fichier_pdf, fichier_excel, debut, fin,
                  nb_alertes, org, datetime.datetime.now().isoformat()))


# ── GENERATEUR PDF ────────────────────────────────────────────────
class PDFGenerator:

    def __init__(self, cfg, db_obj):
        self.cfg = cfg
        self.db  = db_obj
        self.org = db_obj.get_param('organisation_nom', cfg.get('ORG_NOM', 'SIEM Africa'))
        self.lang= cfg.get('LANG', 'fr')
        self.W, self.H = A4
        self.CW = self.W - 4*cm

    def _S(self, name, **kw): return ParagraphStyle(name, **kw)

    def _styles(self):
        return {
            'h1':   self._S('h1', fontName='Helvetica-Bold', fontSize=14, textColor=NOIR,
                            spaceAfter=4, spaceBefore=14, leading=18),
            'h2':   self._S('h2', fontName='Helvetica-Bold', fontSize=11, textColor=SUB,
                            spaceAfter=3, spaceBefore=10, leading=15),
            'body': self._S('bd', fontName='Helvetica', fontSize=10, textColor=SUB,
                            spaceAfter=4, leading=15, alignment=TA_JUSTIFY),
            'th':   self._S('th', fontName='Helvetica-Bold', fontSize=9, textColor=white,
                            spaceAfter=0, leading=12),
            'td':   self._S('td', fontName='Helvetica', fontSize=9, textColor=SUB,
                            spaceAfter=0, leading=13),
            'tdb':  self._S('tdb',fontName='Helvetica-Bold', fontSize=9, textColor=NOIR,
                            spaceAfter=0, leading=13),
            'sm':   self._S('sm', fontName='Helvetica', fontSize=8, textColor=MUTED,
                            spaceAfter=2, leading=12),
            'ctr':  self._S('ctr',fontName='Helvetica', fontSize=9, textColor=SUB,
                            spaceAfter=0, leading=13, alignment=TA_CENTER),
        }

    def P(self, t, s='body'):
        ST = self._styles()
        return Paragraph(t, ST.get(s, ST['body']))

    def SP(self, h=0.25): return Spacer(1, h*cm)
    def HR(self): return HRFlowable(width='100%', thickness=0.5,
                                     color=BORD, spaceAfter=4, spaceBefore=3)

    def tbl(self, headers, rows, widths=None):
        data = [[self.P(h, 'th') for h in headers]] + rows
        t = Table(data, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0),(-1,0), TH_BG),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [white, BGL]),
            ('GRID',          (0,0),(-1,-1), 0.3, BORD),
            ('PADDING',       (0,0),(-1,-1), 6),
            ('VALIGN',        (0,0),(-1,-1), 'TOP'),
            ('LINEBELOW',     (0,0),(-1,0), 0.8, TH_BG),
        ]))
        return t

    def on_page(self, titre_rapport, num_pages=None):
        def _f(canvas, doc):
            canvas.saveState()
            # En-tete
            canvas.setFillColor(TH_BG)
            canvas.rect(0, self.H-1.2*cm, self.W, 1.2*cm, fill=1, stroke=0)
            canvas.setFillColor(white)
            canvas.setFont('Helvetica-Bold', 9)
            canvas.drawString(2*cm, self.H-0.8*cm, f'SIEM Africa  —  {titre_rapport}')
            canvas.setFont('Helvetica', 8)
            canvas.drawRightString(self.W-2*cm, self.H-0.8*cm, self.org)
            # Pied
            canvas.setStrokeColor(BORD)
            canvas.setLineWidth(0.3)
            canvas.line(2*cm, 1.2*cm, self.W-2*cm, 1.2*cm)
            canvas.setFillColor(MUTED)
            canvas.setFont('Helvetica', 8)
            canvas.drawString(2*cm, 0.7*cm, 'SIEM Africa — Document confidentiel')
            canvas.drawRightString(self.W-2*cm, 0.7*cm, f'Page {doc.page}')
            canvas.restoreState()
        return _f

    def _gravite_label(self, g):
        labels = {4:'Critique', 3:'Haute', 2:'Moyenne', 1:'Faible'}
        return labels.get(int(g) if g else 0, 'Inconnue')

    def _score_securite(self, stats):
        """Score de securite 0-100"""
        total = stats['total'] or 1
        tx_res = stats['resolues'] / total
        ratio_crit = stats['critiques'] / total
        score = int(100 * tx_res * (1 - ratio_crit * 0.5))
        return max(0, min(100, score))

    # ── Rapport d'incident ─────────────────────────────────────────
    def rapport_incident(self, alerte_id, output_path):
        alerte = self.db.get_alerte(alerte_id)
        if not alerte:
            log.error(f'Alerte {alerte_id} introuvable')
            return None

        doc = SimpleDocTemplate(output_path, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm)
        titre = f'Rapport d\'incident #{alerte_id}'
        s = []

        # En-tete
        s += [self.SP(0.5)]
        s.append(self.P(f'<b>{titre}</b>', 'h1'))
        s.append(self.HR())
        s.append(self.SP(0.2))

        # Fiche incident
        gravite = alerte.get('gravite', 0)
        gravite_label = self._gravite_label(gravite)
        s.append(self.tbl(['', ''], [
            [self.P('Organisation', 'tdb'),  self.P(self.org, 'td')],
            [self.P('Date de detection', 'tdb'), self.P(str(alerte.get('timestamp_alerte',''))[:16], 'td')],
            [self.P('Gravite', 'tdb'),       self.P(gravite_label, 'tdb')],
            [self.P('Attaque', 'tdb'),       self.P(str(alerte.get('nom_attaque', alerte.get('attaque_nom','—'))), 'td')],
            [self.P('Categorie', 'tdb'),     self.P(str(alerte.get('categorie','—')), 'td')],
            [self.P('IP source', 'tdb'),     self.P(str(alerte.get('ip_source','—')), 'td')],
            [self.P('Pays', 'tdb'),          self.P(str(alerte.get('pays_source','—')), 'td')],
            [self.P('Machine ciblee', 'tdb'),self.P(str(alerte.get('machine_nom','—')), 'td')],
            [self.P('Statut', 'tdb'),        self.P(str(alerte.get('statut','—')), 'td')],
            [self.P('MITRE ID', 'tdb'),      self.P(str(alerte.get('mitre_id','—')), 'td')],
            [self.P('Tactique MITRE', 'tdb'),self.P(str(alerte.get('mitre_tactique','—')), 'td')],
            [self.P('Traite par', 'tdb'),    self.P(str(alerte.get('resolu_par_email','Automatique')), 'td')],
        ], widths=[5*cm, self.CW-5*cm]))

        if alerte.get('action_recommandee'):
            s.append(self.SP())
            s.append(self.P('Action recommandee', 'h2'))
            s.append(self.P(str(alerte['action_recommandee']), 'body'))

        if alerte.get('contre_mesure'):
            s.append(self.SP())
            s.append(self.P('Contre-mesure executee', 'h2'))
            s.append(self.P(str(alerte['contre_mesure']), 'body'))

        if alerte.get('description_wazuh'):
            s.append(self.SP())
            s.append(self.P('Description technique (Wazuh)', 'h2'))
            s.append(self.P(str(alerte['description_wazuh']), 'body'))

        s.append(self.SP(0.5))
        s.append(self.HR())
        s.append(self.P(f'Rapport genere automatiquement par SIEM Africa le {datetime.datetime.now().strftime("%d/%m/%Y a %H:%M")}', 'sm'))

        doc.build(s, onFirstPage=self.on_page(titre), onLaterPages=self.on_page(titre))
        log.info(f'PDF incident genere : {output_path}')
        return output_path

    # ── Rapport periodique (hebdo/trim/annuel/manuel) ──────────────
    def rapport_periodique(self, type_rapport, debut, fin, output_path):
        stats   = self.db.stats_periode(debut, fin)
        alertes = self.db.alertes_periode(debut, fin)
        top_att = self.db.top_attaques(debut, fin)
        top_ips = self.db.top_ips(debut, fin)
        par_jour= self.db.alertes_par_jour(debut, fin)
        score   = self._score_securite(stats)

        titres = {
            'hebdomadaire': 'Rapport Hebdomadaire de Securite',
            'trimestriel':  'Rapport Trimestriel de Securite',
            'annuel':       'Rapport Annuel de Securite',
            'manuel':       'Rapport de Securite',
        }
        titre = titres.get(type_rapport, 'Rapport de Securite')

        doc = SimpleDocTemplate(output_path, pagesize=A4,
            leftMargin=2*cm, rightMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm)
        s = []

        # Titre
        s.append(self.SP(0.5))
        s.append(self.P(f'<b>{titre}</b>', 'h1'))
        s.append(self.P(f'{self.org}  —  du {debut[:10]} au {fin[:10]}', 'sm'))
        s.append(self.HR())
        s.append(self.SP(0.3))

        # Score securite
        score_color = '#27AE60' if score >= 70 else ('#E67E22' if score >= 40 else '#C0392B')
        s.append(self.P(f'<b>Score de securite : <font color="{score_color}">{score}/100</font></b>', 'h2'))
        s.append(self.SP(0.2))

        # Statistiques generales
        s.append(self.P('Statistiques generales', 'h2'))
        s.append(self.tbl(['Indicateur', 'Valeur'], [
            [self.P('Total alertes', 'tdb'),     self.P(str(stats['total']), 'ctr')],
            [self.P('Critiques', 'tdb'),          self.P(str(stats['critiques']), 'ctr')],
            [self.P('Hautes', 'tdb'),             self.P(str(stats['hautes']), 'ctr')],
            [self.P('Moyennes', 'tdb'),           self.P(str(stats['moyennes']), 'ctr')],
            [self.P('Faibles', 'tdb'),            self.P(str(stats['faibles']), 'ctr')],
            [self.P('Resolues', 'tdb'),           self.P(str(stats['resolues']), 'ctr')],
            [self.P('Faux positifs', 'tdb'),      self.P(str(stats['faux_pos']), 'ctr')],
            [self.P('IPs bloquees', 'tdb'),       self.P(str(stats['ips_bloquees']), 'ctr')],
            [self.P('Alertes honeypot', 'tdb'),   self.P(str(stats['honeypot']), 'ctr')],
            [self.P('Alertes correlees', 'tdb'),  self.P(str(stats['correlees']), 'ctr')],
            [self.P('Attaques inconnues', 'tdb'), self.P(str(stats['inconnues']), 'ctr')],
        ], widths=[8*cm, self.CW-8*cm]))
        s.append(self.SP())

        # Top attaques
        if top_att:
            s.append(self.P('Top attaques detectees', 'h2'))
            rows = [[
                self.P(str(a.get('nom_attaque','—')), 'td'),
                self.P(str(a.get('categorie','—')), 'td'),
                self.P(str(a.get('nb',0)), 'ctr'),
                self.P(self._gravite_label(a.get('gravite_max',1)), 'td'),
                self.P(str(a.get('nb_ips',0)), 'ctr'),
            ] for a in top_att]
            s.append(self.tbl(
                ['Attaque', 'Categorie', 'Nb', 'Gravite max', 'IPs uniques'],
                rows,
                widths=[4.5*cm, 3*cm, 1.5*cm, 2.5*cm, 2.5*cm]
            ))
            s.append(self.SP())

        # Top IPs
        if top_ips:
            s.append(self.P('Top IPs sources', 'h2'))
            rows = [[
                self.P(str(i.get('ip_source','—')), 'td'),
                self.P(str(i.get('pays_source','—')), 'td'),
                self.P(str(i.get('nb',0)), 'ctr'),
                self.P(self._gravite_label(i.get('gravite_max',1)), 'td'),
            ] for i in top_ips]
            s.append(self.tbl(
                ['IP Source', 'Pays', 'Nb alertes', 'Gravite max'],
                rows,
                widths=[4*cm, 4*cm, 3*cm, self.CW-11*cm]
            ))
            s.append(self.SP())

        # Evolution par jour
        if par_jour:
            s.append(self.P('Evolution quotidienne', 'h2'))
            rows = [[
                self.P(str(j.get('jour','—')), 'td'),
                self.P(str(j.get('nb',0)), 'ctr'),
                self.P(str(j.get('critiques',0)), 'ctr'),
            ] for j in par_jour[-14:]]  # 14 derniers jours max
            s.append(self.tbl(
                ['Date', 'Total alertes', 'Critiques'],
                rows,
                widths=[4*cm, 4*cm, self.CW-8*cm]
            ))
            s.append(self.SP())

        # Liste des alertes (si pas trop long)
        if alertes and len(alertes) <= 100:
            s.append(PageBreak())
            s.append(self.P('Detail des alertes', 'h2'))
            rows = [[
                self.P(str(a.get('timestamp_alerte',''))[:16], 'td'),
                self.P(str(a.get('nom_attaque','—'))[:50], 'td'),
                self.P(str(a.get('ip_source','—')), 'td'),
                self.P(str(a.get('pays_source','—')), 'td'),
                self.P(self._gravite_label(a.get('gravite',1)), 'td'),
                self.P(str(a.get('statut','—')), 'td'),
            ] for a in alertes]
            s.append(self.tbl(
                ['Date/Heure', 'Attaque', 'IP', 'Pays', 'Gravite', 'Statut'],
                rows,
                widths=[3*cm, 4.5*cm, 2.5*cm, 2*cm, 2*cm, 2*cm]
            ))

        s.append(self.SP(0.5))
        s.append(self.HR())
        s.append(self.P(
            f'Rapport genere automatiquement par SIEM Africa le '
            f'{datetime.datetime.now().strftime("%d/%m/%Y a %H:%M")} — '
            f'github.com/luciesys/SIEM-AFRICA', 'sm'))

        doc.build(s, onFirstPage=self.on_page(titre), onLaterPages=self.on_page(titre))
        log.info(f'PDF {type_rapport} genere : {output_path}')
        return output_path


# ── GENERATEUR EXCEL ──────────────────────────────────────────────
class ExcelGenerator:

    def __init__(self, cfg, db_obj):
        self.cfg = cfg
        self.db  = db_obj
        self.org = db_obj.get_param('organisation_nom', cfg.get('ORG_NOM', 'SIEM Africa'))

    def _style_header(self, ws, row, cols, fill_color='1A1A2E'):
        fill = PatternFill('solid', fgColor=fill_color)
        font = Font(bold=True, color='FFFFFF', size=11)
        border= Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col in range(1, cols+1):
            cell = ws.cell(row=row, column=col)
            cell.fill   = fill
            cell.font   = font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _style_data(self, ws, row, cols, alt=False):
        fill   = PatternFill('solid', fgColor='F8F8F8' if alt else 'FFFFFF')
        border = Border(
            left=Side(style='thin',color='CCCCCC'),
            right=Side(style='thin',color='CCCCCC'),
            top=Side(style='thin',color='CCCCCC'),
            bottom=Side(style='thin',color='CCCCCC')
        )
        for col in range(1, cols+1):
            c = ws.cell(row=row, column=col)
            c.fill   = fill
            c.border = border
            c.alignment = Alignment(vertical='center', wrap_text=True)

    def _gravite_fill(self, g):
        colors = {4:'C0392B', 3:'E67E22', 2:'F1C40F', 1:'27AE60'}
        return colors.get(int(g) if g else 0, 'CCCCCC')

    # ── Rapport incident Excel ─────────────────────────────────────
    def rapport_incident(self, alerte_id, output_path):
        alerte = self.db.get_alerte(alerte_id)
        if not alerte:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = f'Incident #{alerte_id}'
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 55

        # Titre
        ws.merge_cells('A1:B1')
        ws['A1'] = f'RAPPORT D\'INCIDENT #{alerte_id} — {self.org}'
        ws['A1'].font      = Font(bold=True, size=14, color='1A1A2E')
        ws['A1'].fill      = PatternFill('solid', fgColor='F0F0F0')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        donnees = [
            ('Date detection',    str(alerte.get('timestamp_alerte',''))[:16]),
            ('Attaque',           str(alerte.get('nom_attaque', alerte.get('attaque_nom','—')))),
            ('Categorie',         str(alerte.get('categorie','—'))),
            ('Gravite',           f"{alerte.get('gravite',0)} — {['','Faible','Moyenne','Haute','Critique'].get(int(alerte.get('gravite',0)),'?')}" if alerte.get('gravite') else '—'),
            ('IP source',         str(alerte.get('ip_source','—'))),
            ('Pays',              str(alerte.get('pays_source','—'))),
            ('Machine ciblee',    str(alerte.get('machine_nom','—'))),
            ('Statut',            str(alerte.get('statut','—'))),
            ('MITRE ID',          str(alerte.get('mitre_id','—'))),
            ('Tactique MITRE',    str(alerte.get('mitre_tactique','—'))),
            ('Action prise',      str(alerte.get('action_recommandee','—'))),
            ('Contre-mesure',     str(alerte.get('contre_mesure','—'))),
            ('Traite par',        str(alerte.get('resolu_par_email','Automatique'))),
            ('Description Wazuh', str(alerte.get('description_wazuh','—'))),
        ]

        for i, (label, valeur) in enumerate(donnees, start=2):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
            ws.cell(row=i, column=2, value=valeur).font = Font(size=10)
            ws.cell(row=i, column=1).fill = PatternFill('solid', fgColor='F8F8F8' if i%2==0 else 'FFFFFF')
            ws.row_dimensions[i].height = 20

        ws['A2'].fill = PatternFill('solid', fgColor='F0F0F0')
        ws['B2'].fill = PatternFill('solid', fgColor='F0F0F0')

        wb.save(output_path)
        log.info(f'Excel incident genere : {output_path}')
        return output_path

    # ── Rapport periodique Excel ───────────────────────────────────
    def rapport_periodique(self, type_rapport, debut, fin, output_path):
        stats   = self.db.stats_periode(debut, fin)
        alertes = self.db.alertes_periode(debut, fin)
        top_att = self.db.top_attaques(debut, fin)
        top_ips = self.db.top_ips(debut, fin)
        par_jour= self.db.alertes_par_jour(debut, fin)

        wb = Workbook()

        # ── Onglet Resume ──────────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'Resume'
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20

        ws1.merge_cells('A1:B1')
        titres = {'hebdomadaire':'Hebdomadaire','trimestriel':'Trimestriel','annuel':'Annuel','manuel':'Rapport'}
        ws1['A1'] = f'RAPPORT {titres.get(type_rapport,"").upper()} — {self.org}'
        ws1['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
        ws1['A1'].fill      = PatternFill('solid', fgColor='1A1A2E')
        ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[1].height = 32

        ws1['A2'] = f'Periode : {debut[:10]} au {fin[:10]}'
        ws1['A2'].font = Font(italic=True, size=10, color='666666')
        ws1.merge_cells('A2:B2')

        score = int(100 * (stats['resolues']/(stats['total'] or 1)) * (1 - stats['critiques']/(stats['total'] or 1) * 0.5))
        score = max(0, min(100, score))

        indicateurs = [
            ('Total alertes',     stats['total']),
            ('Critiques',         stats['critiques']),
            ('Hautes',            stats['hautes']),
            ('Moyennes',          stats['moyennes']),
            ('Faibles',           stats['faibles']),
            ('Resolues',          stats['resolues']),
            ('Faux positifs',     stats['faux_pos']),
            ('IPs bloquees',      stats['ips_bloquees']),
            ('Alertes honeypot',  stats['honeypot']),
            ('Alertes correlees', stats['correlees']),
            ('Attaques inconnues',stats['inconnues']),
            ('Score securite',    f'{score}/100'),
        ]
        for i, (label, val) in enumerate(indicateurs, start=4):
            ws1.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
            ws1.cell(row=i, column=2, value=val).font = Font(size=10)
            ws1.cell(row=i, column=2).alignment = Alignment(horizontal='center')
            self._style_data(ws1, i, 2, alt=i%2==0)
            ws1.row_dimensions[i].height = 20

        # Couleur score
        score_fill = '27AE60' if score>=70 else ('E67E22' if score>=40 else 'C0392B')
        ws1[f'B{3+len(indicateurs)}'].fill = PatternFill('solid', fgColor=score_fill)
        ws1[f'B{3+len(indicateurs)}'].font = Font(bold=True, color='FFFFFF', size=10)

        # ── Onglet Alertes ─────────────────────────────────────────
        ws2 = wb.create_sheet('Alertes')
        headers = ['Date/Heure', 'Attaque', 'Categorie', 'Gravite', 'IP Source', 'Pays', 'Machine', 'Statut', 'MITRE']
        widths  = [18, 30, 15, 10, 15, 12, 15, 12, 12]
        for i, (h, w) in enumerate(zip(headers, widths), start=1):
            ws2.cell(row=1, column=i, value=h)
            ws2.column_dimensions[get_column_letter(i)].width = w
        self._style_header(ws2, 1, len(headers))
        ws2.row_dimensions[1].height = 22
        ws2.freeze_panes = 'A2'

        for ri, a in enumerate(alertes, start=2):
            g = int(a.get('gravite', 1))
            g_label = ['','Faible','Moyenne','Haute','Critique'][g] if 1<=g<=4 else '?'
            vals = [
                str(a.get('timestamp_alerte',''))[:16],
                str(a.get('nom_attaque','—')),
                str(a.get('categorie','—')),
                g_label,
                str(a.get('ip_source','—')),
                str(a.get('pays_source','—')),
                str(a.get('machine_nom','—')),
                str(a.get('statut','—')),
                str(a.get('mitre_id','—')),
            ]
            for ci, v in enumerate(vals, start=1):
                ws2.cell(row=ri, column=ci, value=v).font = Font(size=9)
            self._style_data(ws2, ri, len(headers), alt=ri%2==0)
            ws2.row_dimensions[ri].height = 18
            # Couleur gravite
            fill_g = PatternFill('solid', fgColor=self._gravite_fill(g)+'33')
            ws2.cell(row=ri, column=4).fill = fill_g

        # ── Onglet Top Attaques ────────────────────────────────────
        ws3 = wb.create_sheet('Top Attaques')
        h3 = ['Attaque', 'Categorie', 'Nombre', 'Gravite max', 'IPs uniques']
        w3 = [35, 18, 10, 14, 12]
        for i, (h, w) in enumerate(zip(h3, w3), start=1):
            ws3.cell(row=1, column=i, value=h)
            ws3.column_dimensions[get_column_letter(i)].width = w
        self._style_header(ws3, 1, len(h3))
        ws3.row_dimensions[1].height = 22
        for ri, a in enumerate(top_att, start=2):
            g = int(a.get('gravite_max',1))
            vals = [a.get('nom_attaque','—'), a.get('categorie','—'), a.get('nb',0),
                    ['','Faible','Moyenne','Haute','Critique'][g] if 1<=g<=4 else '?',
                    a.get('nb_ips',0)]
            for ci, v in enumerate(vals, start=1):
                ws3.cell(row=ri, column=ci, value=v).font = Font(size=9)
            self._style_data(ws3, ri, len(h3), alt=ri%2==0)
            ws3.row_dimensions[ri].height = 18

        # ── Onglet Evolution ───────────────────────────────────────
        ws4 = wb.create_sheet('Evolution')
        h4 = ['Date', 'Total alertes', 'Critiques']
        w4 = [15, 16, 12]
        for i, (h, w) in enumerate(zip(h4, w4), start=1):
            ws4.cell(row=1, column=i, value=h)
            ws4.column_dimensions[get_column_letter(i)].width = w
        self._style_header(ws4, 1, len(h4))
        for ri, j in enumerate(par_jour, start=2):
            ws4.cell(row=ri, column=1, value=j.get('jour',''))
            ws4.cell(row=ri, column=2, value=j.get('nb',0))
            ws4.cell(row=ri, column=3, value=j.get('critiques',0))
            self._style_data(ws4, ri, len(h4), alt=ri%2==0)
            ws4.row_dimensions[ri].height = 18

        wb.save(output_path)
        log.info(f'Excel {type_rapport} genere : {output_path}')
        return output_path


# ── FONCTION PRINCIPALE ───────────────────────────────────────────
def generer_rapport(type_rapport, alerte_id=None, debut=None, fin=None):
    """
    type_rapport : 'incident', 'hebdomadaire', 'trimestriel', 'annuel', 'manuel'
    alerte_id    : pour type='incident' uniquement
    debut/fin    : pour les autres types (format ISO)
    """
    cfg = load_env()
    reports_dir = cfg.get('REPORTS_DIR', REPORTS_DIR)
    os.makedirs(reports_dir, exist_ok=True)

    db = DB(cfg['DB_PATH'])

    # Calculer la periode automatiquement si non fournie
    now = datetime.datetime.now()
    if type_rapport == 'hebdomadaire' and not debut:
        lundi = now - datetime.timedelta(days=now.weekday()+7)
        debut = lundi.strftime('%Y-%m-%d 00:00:00')
        fin   = (lundi + datetime.timedelta(days=6)).strftime('%Y-%m-%d 23:59:59')
    elif type_rapport == 'trimestriel' and not debut:
        trimestre = (now.month - 1) // 3
        mois_debut= trimestre * 3 + 1
        if mois_debut == 1:
            debut = f'{now.year-1}-10-01 00:00:00'
            fin   = f'{now.year-1}-12-31 23:59:59'
        else:
            debut = f'{now.year}-{mois_debut-3:02d}-01 00:00:00'
            fin   = (now.replace(day=1) - datetime.timedelta(days=1)).strftime('%Y-%m-%d 23:59:59')
    elif type_rapport == 'annuel' and not debut:
        debut = f'{now.year-1}-01-01 00:00:00'
        fin   = f'{now.year-1}-12-31 23:59:59'

    # Nom des fichiers
    ts     = now.strftime('%Y%m%d_%H%M%S')
    suffix = f'incident_{alerte_id}' if type_rapport == 'incident' else type_rapport
    pdf_path  = os.path.join(reports_dir, f'siem_africa_{suffix}_{ts}.pdf')
    xlsx_path = os.path.join(reports_dir, f'siem_africa_{suffix}_{ts}.xlsx')

    pdf_gen   = PDFGenerator(cfg, db) if PDF_OK else None
    excel_gen = ExcelGenerator(cfg, db) if EXCEL_OK else None

    pdf_ok  = None
    xlsx_ok = None

    if type_rapport == 'incident':
        if pdf_gen:
            pdf_ok = pdf_gen.rapport_incident(alerte_id, pdf_path)
        if excel_gen:
            xlsx_ok = excel_gen.rapport_incident(alerte_id, xlsx_path)
    else:
        if pdf_gen:
            pdf_ok = pdf_gen.rapport_periodique(type_rapport, debut, fin, pdf_path)
        if excel_gen:
            xlsx_ok = excel_gen.rapport_periodique(type_rapport, debut, fin, xlsx_path)

    # Sauvegarder en base
    alertes = db.alertes_periode(debut or '', fin or '') if type_rapport != 'incident' else []
    nb      = db.get_alerte(alerte_id) if type_rapport == 'incident' else {}
    db.sauver_rapport(
        type_rapport,
        pdf_ok or '',
        xlsx_ok or '',
        debut or now.isoformat(),
        fin or now.isoformat(),
        len(alertes) if alertes else (1 if type_rapport == 'incident' else 0),
        db.get_param('organisation_nom', cfg.get('ORG_NOM', ''))
    )

    return pdf_ok, xlsx_ok


if __name__ == '__main__':
    import argparse
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s [%(levelname)s] %(message)s')

    parser = argparse.ArgumentParser(description='SIEM Africa — Generateur de rapports')
    parser.add_argument('type', choices=['incident','hebdomadaire','trimestriel','annuel','manuel'],
                        help='Type de rapport')
    parser.add_argument('--alerte-id', type=int, help='ID alerte (pour type=incident)')
    parser.add_argument('--debut', help='Date debut ISO (pour manuel)')
    parser.add_argument('--fin',   help='Date fin ISO (pour manuel)')
    args = parser.parse_args()

    pdf, xlsx = generer_rapport(
        args.type,
        alerte_id=args.alerte_id,
        debut=args.debut,
        fin=args.fin
    )
    if pdf:   print(f'PDF   : {pdf}')
    if xlsx:  print(f'Excel : {xlsx}')
